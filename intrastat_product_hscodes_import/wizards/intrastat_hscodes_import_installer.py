import os

from openpyxl import load_workbook

from odoo import fields, models, tools
from odoo.exceptions import MissingError, UserError
from odoo.fields import Domain

UOM_MAPPING = {
    "p/st": "intrastat_unit_pce",
    "100 p/st": "intrastat_unit_100pce",
    "1 000 p/st": "intrastat_unit_1000pce",
    "l alc. 100 %": "intrastat_unit_l_alc_100_pct",
    "kg 90 % sdt": "intrastat_unit_kg_90_pct_sdt",
    "m²": "intrastat_unit_m2",
    "m³": "intrastat_unit_m3",
    "1 000 m³": "intrastat_unit_1000m3",
    "1 000 kWh": "intrastat_unit_1000_kWh",
}


class IntrastatNomenclatureCodesImportInstaller(models.TransientModel):
    _name = "intrastat.hscodes.import.installer"
    _description = "Intrastat Nomenclature Import Installer"

    language_id = fields.Many2one("res.lang", string="Language")

    @tools.ormcache("name")
    def _get_intrastat_unit(self, name):
        return self.env["intrastat.unit"].search(Domain("name", "=", name), limit=1).id

    def _read_excel_sheet(self, file_path, sheet_index=0):
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        header = [
            cell.value.strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]
        return sheet, header

    def _process_uom(self):
        uom_path = os.path.join(
            tools.misc.file_path(
                "intrastat_product_hscodes_import/data/measures/units_of_measure.xlsx"
            )
        )
        uom_sheet, uom_header = self._read_excel_sheet(uom_path)
        uom_code_index = uom_header.index("CNKEY")
        uom_label_index = uom_header.index("SU")
        uom_map = {}
        for row in uom_sheet.iter_rows(min_row=2, values_only=True):
            code = str(row[uom_code_index] or "").strip().replace(" ", "")
            label = str(row[uom_label_index] or "").replace("\xa0", " ").strip()
            if code and label and label != "-":
                uom_map[code] = label
        return uom_map

    def _import_local_codes(self, uom_map):
        code_obj = self.env["hs.code"].with_context(active_test=False)
        codes_path = os.path.join(
            tools.misc.file_path(
                "intrastat_product_hscodes_import/data/codes/declarable_codes.xlsx"
            )
        )
        codes_sheet, codes_header = self._read_excel_sheet(codes_path)
        goods_code_index = codes_header.index("Goods code")
        vals_list = []
        seen_codes = set()
        for row in codes_sheet.iter_rows(min_row=2, values_only=True):
            raw_code = str(row[goods_code_index]).strip().replace(" ", "")
            code_value = raw_code[:8]
            existing = code_obj.search(Domain("local_code", "=", code_value), limit=1)
            if existing:
                continue
            if not code_value or not code_value.isdigit():
                continue
            if code_value in seen_codes:
                continue
            seen_codes.add(code_value)
            vals = {"local_code": code_value}
            if raw_code in uom_map:
                iu = uom_map[raw_code]
                iu_unit_id = self._get_mapped_uom_id(iu)
                if not iu_unit_id:
                    raise UserError(self.env._("Unit not found: '%s'", iu))
                vals["intrastat_unit_id"] = iu_unit_id
            vals_list.append(vals)
        if vals_list:
            code_obj.create(vals_list)
        return code_obj

    def _get_mapped_uom_id(self, uom_label):
        if uom_label in UOM_MAPPING:
            return self.env.ref(f"intrastat_product.{UOM_MAPPING[uom_label]}").id
        else:
            return self._get_intrastat_unit(uom_label)

    def _update_existing_code(self, existing_code, vals):
        if "intrastat_unit_id" in vals and not existing_code.intrastat_unit_id:
            existing_code.write({"intrastat_unit_id": vals["intrastat_unit_id"]})

    def _import_country_specific_descriptions(self, code_obj):
        short_code = self.language_id.code.split("_")[0]
        filename = f"{short_code}.xlsx"
        file_path = tools.misc.file_path(
            f"intrastat_product_hscodes_import/data/countries/{filename}"
        )
        if not file_path:
            raise MissingError(
                self.env._("Missing language file for code '%s'", short_code)
            )
        sheet, header = self._read_excel_sheet(file_path)
        indent_index = header.index("Indent")
        description_index = header.index("Description")
        goods_code_index = header.index("Goods code")
        all_codes = code_obj.search_read(Domain.TRUE, ["id", "local_code"])
        code_map = {
            rec["local_code"].replace(" ", "")[:8]: rec["id"] for rec in all_codes
        }
        stack = []
        last_indent = 0
        prev_description = ""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_code = row[goods_code_index]
            code_value = str(raw_code).strip().replace(" ", "")[:8]
            code_id = code_map.get(code_value)
            if not code_id:
                continue
            description = row[description_index].strip()
            indent_raw = row[indent_index] or ""
            indent = indent_raw.strip().count("-")
            if indent > last_indent:
                stack.append(prev_description)
            elif indent < last_indent:
                stack = stack[:indent]
            full_description = " > ".join(stack + [description])
            prev_description = description
            last_indent = indent
            code_obj.browse(code_id).with_context(lang=self.language_id.code).write(
                {"description": full_description}
            )

    def do_import(self):
        uom_map = self._process_uom()
        code_obj = self._import_local_codes(uom_map)
        if self.language_id:
            self._import_country_specific_descriptions(code_obj)
