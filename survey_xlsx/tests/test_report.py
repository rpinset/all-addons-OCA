# Copyright 2022 Creu Blanca
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
import io
import logging
from datetime import date

import freezegun
import openpyxl

from odoo.addons.survey.tests import common

_logger = logging.getLogger(__name__)


@freezegun.freeze_time("2022-04-26")
class TestReport(common.TestSurveyCommon):
    def setUp(self):
        super().setUp()
        self.question_date = (
            self.env["survey.question"]
            .with_user(self.survey_manager)
            .create(
                {
                    "title": "Test Date",
                    "survey_id": self.survey.id,
                    "sequence": 4,
                    "question_type": "date",
                }
            )
        )
        self.suggested_question = self._add_question(
            page=self.question_date.page_id,
            name="Test Suggested",
            qtype="simple_choice",
            labels=[
                {"value": "FIRST", "suggested_answer_id": True},
                {"value": "SECOND"},
            ],
            survey_id=self.survey.id,
            sequence=5,
        )

        answer = self._add_answer(
            self.survey,
            self.survey_manager.partner_id,
            email=self.survey_manager.partner_id.email,
        )
        self._add_answer_line(self.question_ft, answer, "FIRST ANSWER")
        self._add_answer_line(self.question_num, answer, 1)
        self._add_answer_line(self.question_date, answer, date.today())
        self._add_answer_line(
            self.suggested_question,
            answer,
            self.suggested_question.suggested_answer_ids[0]["id"],
        )
        answer._mark_done()
        answer2 = self._add_answer(self.survey, False, email="public2@example.com")
        self._add_answer_line(self.question_ft, answer2, "SECOND ANSWER")
        self._add_answer_line(self.question_num, answer2, 2)
        self._add_answer_line(
            self.question_date, answer2, False, skipped=True, answer_type=False
        )
        self._add_answer_line(
            self.suggested_question,
            answer2,
            self.suggested_question.suggested_answer_ids[1]["id"],
        )
        answer2._mark_done()

    def test_report(self):
        report = self.env.ref("survey_xlsx.report_survey_xlsx")
        self.assertEqual(report.report_type, "xlsx")
        rep = self.env["ir.actions.report"]._render(report, self.survey.ids, {})
        report_file = io.BytesIO(rep[0])
        wb = openpyxl.load_workbook(report_file)
        sheet = wb.worksheets[0]
        self.assertEqual(sheet.cell(2, 3).value, "FIRST ANSWER")
        self.assertEqual(sheet.cell(3, 3).value, "SECOND ANSWER")
        self.assertEqual(sheet.cell(2, 4).value, 1)
        self.assertEqual(sheet.cell(3, 4).value, 2)
        self.assertEqual(sheet.cell(2, 6).value, "2022-04-26")
        self.assertFalse(sheet.cell(3, 6).value)
        self.assertEqual(sheet.cell(2, 7).value, "FIRST")
        self.assertEqual(sheet.cell(3, 7).value, "SECOND")
