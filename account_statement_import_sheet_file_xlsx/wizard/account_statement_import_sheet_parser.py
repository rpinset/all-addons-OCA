# Copyright 2019 ForgeFlow, S.L.
# Copyright 2020 CorporateHub (https://corporatehub.eu)
# Copyright 2025 Jacques-Etienne Baudoux (BCIM) <je@bcim.be>
# Copyright 2026 Andrii Kompaniiets - Tecnativa
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).

from datetime import datetime
from io import BytesIO

import openpyxl

from odoo import api, models


class AccountStatementImportSheetParser(models.TransientModel):
    _inherit = "account.statement.import.sheet.parser"

    def _get_mimetype_sheet_type_map(self):
        """Return supported MIME types and their parser suffixes."""
        res = super()._get_mimetype_sheet_type_map()
        xlsx_key = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        res.update(
            {
                xlsx_key: "xlsx",
            }
        )
        return res

    @api.model
    def parse_header_xlsx(self, sheet, mapping):
        """
        Return the XLSX header row after applying configured skips and offset.
        """
        if mapping.no_header:
            return []
        row_index = max(mapping.header_lines_skip_count, 1)
        row = next(
            sheet.iter_rows(
                min_row=row_index,
                max_row=row_index,
                values_only=True,
            )
        )
        header = [str(value or "").strip() for value in row]
        if mapping.offset_column:
            header = header[mapping.offset_column :]
        return header

    def _parse_lines_xlsx(self, mapping, data_file, currency_code):
        """
        Parse XLSX content into normalized statement line dictionaries.
        :return: Parsed statement lines ready for transaction conversion.
        """
        columns = dict()
        workbook = openpyxl.load_workbook(
            filename=BytesIO(data_file),
            data_only=True,
        )
        xlsx = workbook.active
        header = self.parse_header_xlsx(xlsx, mapping)
        for column_name in self._get_column_names():
            columns[column_name] = self._get_column_indexes(
                header, column_name, mapping
            )
        numrows = xlsx.max_row

        label_line = mapping.header_lines_skip_count
        footer_line = numrows - mapping.footer_lines_skip_count
        rows = range(label_line + 1, footer_line + 1)
        rows_values = self._get_xlsx_row_values(mapping, xlsx, rows, label_line)
        data = rows_values, label_line, footer_line
        return self._parse_rows(mapping, currency_code, data, columns)

    def _get_xlsx_row_values(self, mapping, xlsx, rows, label_line):
        """
        Return row values from an XLSX sheet with date cells formatted.
        """
        parsed_rows = []
        for _, row in enumerate(rows, label_line):
            values = []
            for col_index in range(mapping.offset_column + 1, xlsx.max_column + 1):
                cell_value = xlsx.cell(row=row, column=col_index).value
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime(mapping.timestamp_format)
                values.append(str(cell_value))
            parsed_rows.append(values)
        return parsed_rows
